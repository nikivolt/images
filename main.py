import openpyxl

inv_file = openpyxl.load_workbook("web_import.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
quantity_per_supplier = {}
quantity_less_than_10 = {}

# so I check the max value of rows
# print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    # row 2 - column B, Partnumber
    part_number = product_list.cell(product_row, 2).value
    # row 3 - column C, NameOfProduct
    name_of_product = product_list.cell(product_row, 3).value
    # row 5 - column E, Category
    # we don't add a .value when we want to change the content # .value
    category_of_product = product_list.cell(product_row, 5)  # look the row above
    # row 10 - column J, Brand
    supplier_name = product_list.cell(product_row, 10).value  # so extract/shows the content in the cell
    # row 13 - column M, Quantity
    supplier_quantity = product_list.cell(product_row, 13).value

    # calculate number of suppliers
    if supplier_name in products_per_supplier:
        # get values of dict (recommended way)
        current_num_products = products_per_supplier.get(supplier_name)
        # another way to get values of dict
        # current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculate the quantity of supplier
    if supplier_name in quantity_per_supplier:
        current_quantity = quantity_per_supplier.get(supplier_name)
        quantity_per_supplier[supplier_name] = current_quantity + supplier_quantity
    else:
        quantity_per_supplier[supplier_name] = supplier_quantity

    # shows when quantity of product is less than 10
    if supplier_quantity < 10:
        quantity_less_than_10[part_number] = supplier_quantity

print(products_per_supplier)
print(quantity_per_supplier)
print(quantity_less_than_10)
