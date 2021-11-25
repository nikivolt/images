import openpyxl


# product list for easy name changer
replace_names = {"монт.Prisma": "монтажна Prisma",
                 "хориз.за": "хоризонтална за",
                 "верт.за": "вертикална за",
                 "универс.елем.20": "универсален елемент 20",
                 "верт.фикс.NSX/CVS/VIGI/INS": "вертикално фиксиране NSX/CVS/VIGI/INS"}


def change_names(excel_path):
    inv_file = openpyxl.load_workbook(excel_path)
    product_list = inv_file.active

    for row in range(2, product_list.max_row + 1):
        product_name = str(product_list.cell(row, 2).value)
        name_separator = str(product_name.split())

        # print(name_separator)

        for word, initial in replace_names.items():
            for separator in name_separator:
                if separator in word:
                    product_name = str(product_name.replace(word, initial))
                    product_list.cell(row, 2).value = product_name
        print(product_name)

    inv_file.save("web_import_edited.xlsx")


if __name__ == "__main__":
    change_names("web_import.xlsx")
