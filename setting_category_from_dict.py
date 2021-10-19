import openpyxl

# product class for easy products management
class Category:
    def __init__(self, name, words, required_count=None):
        self.name = name  # name of the category
        self.words = words  # words to be detected
        self.required_count = required_count or len(words)


CATEGORIES = [
    Category('Автоматизация и контрол / Моторни защити', ['моторна', 'защита']),
    Category('Електроинсталационни материали / Контакти / Вграден монтаж', ['контакт', 'открит', 'монтаж'], 2)
]


def update_categories(path):
    inv_file = openpyxl.load_workbook(path)
    product_list = inv_file['Sheet1']

    for product_row in range(2, product_list.max_row + 1):
        # column C (3), NameOfProduct
        product_name = product_list.cell(product_row, 3).value
        # matching dublicated words
        product_name_words = [word.lower() for word in product_name.split()]
        """
        # not matching dublicated words
        product_name_words = set([word.lower() for word in product_name.split()])
        """
        # column D (5), Category
        product_category = product_list.cell(product_row, 5)
        # clear value in cell
        product_category.value = ''

        for category in CATEGORIES:
            count = sum(word in product_name_words for word in category.words)
            print('Found %d/%d matches in %s.' % (count, len(category.words), product_name))
            if count >= category.required_count:
                print(f'The name of product is {product_name}')
                product_category.value = category.name
                break

    inv_file.save(path)


if __name__ == '__main__':
    update_categories('web_import.xlsx')