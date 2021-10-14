import openpyxl


# Product class for easy products management
class Product:
    def __init__(self, productName, category, count):
        self.productName = productName
        self.category = category
        self.count = count


inv_file = openpyxl.load_workbook("web_import.xlsx")
product_list = inv_file["Sheet1"]

dicts = [
    Product(["моторна", "защита"], "Автоматизация и контрол / Моторни защити", 2),
    Product(["контакт", "вграден", "монтаж"], "Електроинсталационни материали / Контакти / Вграден монтаж", 3),
    Product(["филтър", "синусов"], "тест", 2)
]

for product_row in range(2, product_list.max_row + 1):
    # row 3 - column C, NameOfProduct
    name_of_product = product_list.cell(product_row, 3).value
    category_of_product = product_list.cell(product_row, 5)
    divided_words = name_of_product.split()

    for dict in dicts:
        counter = 0
        for w in divided_words:
            for z in dict.productName:
                # match = 0
                while z.lower() in w.lower():
                    # match += 1
                    if z.lower() == w.lower():
                        counter += 1
                        print(counter)
                        if counter <= dict.count:
                            print("equal")

                        if counter >= dict.count:
                        # if match >= dict.count:
                            category_of_product.value = dict.category
                            print(name_of_product)
                            print("-------------------")
                            break
                    else:
                        print("not equal")
                        break

inv_file.save("web_import.xlsx")
