import openpyxl

inv_file = openpyxl.load_workbook("web_import.xlsx")
product_list = inv_file["Sheet1"]

defining_words = ["моторна", "защита"]

counter = 0

for product_row in range(2, product_list.max_row + 1):
    # row 3 - column C, NameOfProduct
    name_of_product = product_list.cell(product_row, 3).value
    divided_words = name_of_product.split()

    for word in defining_words:
        if word in name_of_product:
            counter = counter + 1
            print(name_of_product)

print(counter)
